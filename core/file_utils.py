import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import DEFAULT_XLSX_FILENAME


def criar_diretorio_se_nao_existir(caminho):
    """Cria um diretório se ele não existir."""
    if not os.path.exists(caminho):
        os.makedirs(caminho)


def mover_arquivo(origem, destino_dir):
    """Move um arquivo para o diretório de destino, criando o diretório se necessário."""
    criar_diretorio_se_nao_existir(destino_dir)
    destino = os.path.join(destino_dir, os.path.basename(origem))

    try:
        shutil.move(origem, destino)
        return True
    except Exception as e:
        print(f"Erro ao mover arquivo {origem}: {e}")
        return False


def salvar_em_xlsx(diretorio, dados, nome_arquivo=DEFAULT_XLSX_FILENAME):
    """Salva os dados em um arquivo Excel com formatação específica."""
    try:
        caminho_completo = os.path.join(diretorio, nome_arquivo)

        # Verifica se o arquivo já existe
        if os.path.exists(caminho_completo):
            wb = load_workbook(caminho_completo)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Remove planilha padrão

        # Definir estilos de borda
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Cores pré-definidas
        amarelo_titulo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cinza_cabecalho = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        azul_linha = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        amarelo_duplicado = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        vermelho_novo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        # Fonte padrão Arial
        fonte_normal = Font(name='Arial', size=12)
        fonte_cabecalho = Font(name='Arial', size=12, bold=True)
        fonte_titulo = Font(name='Arial', size=14, bold=True)

        # Formatos numéricos
        formato_inteiro = '0'
        formato_moeda = '"R$" #,##0.00'  # Formato customizado para moeda brasileira
        formato_data = 'dd/mm/yyyy'
        formato_texto = '@'  # Formato de texto no Excel

        # Processar cada PO
        for po_num, linhas in dados.items():
            sheet_name = f"PO_{po_num}"

            # Verifica se a planilha já existe
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # Cria nova planilha com formatação inicial
                ws = wb.create_sheet(sheet_name)

                # Cabeçalho do PO (linha 1)
                ws.merge_cells("A1:E1")
                ws["A1"] = f"PO_{po_num}"
                ws["A1"].font = fonte_titulo
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A1"].fill = amarelo_titulo

                # Cabeçalho "Descrição de Serviços" (linha 2)
                ws.merge_cells("A2:E2")
                ws["A2"] = "DESCRIÇÃO DE SERVIÇOS"
                ws["A2"].font = fonte_titulo
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A2"].fill = cinza_cabecalho

                # Cabeçalhos das colunas (linha 3)
                headers = ["N. Nota", "Data", "Linha da PO", "Descrição", "Valor"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_num, value=header)
                    cell.font = fonte_cabecalho
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = azul_linha

                # Ajuste de altura das linhas
                ws.row_dimensions[1].height = 30
                # Dados Empresas
                ws.row_dimensions[2].height = 30
                ws.row_dimensions[3].height = 20

                # Ajuste de largura das colunas
                ws.column_dimensions["A"].width = 20  # N. Nota
                ws.column_dimensions["B"].width = 20  # Data
                ws.column_dimensions["C"].width = 20  # Linha da PO
                ws.column_dimensions["D"].width = 50  # Descrição
                ws.column_dimensions["E"].width = 20  # Valor

            # Preparar dados para inserção
            novos_dados = []
            for linha in linhas:
                novo_item = [
                    linha.get("nota", ""),
                    linha.get("data_emissao", ""),
                    linha.get("linha", ""),
                    linha.get("descricao", ""),
                    linha.get("valor", 0),
                ]
                novos_dados.append(novo_item)

            # Verificar duplicatas e inserir novos dados
            for novo_item in novos_dados:
                # Verifica se já existe uma linha igual
                duplicata = False
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if row == tuple(novo_item):
                        duplicata = True
                        # Marca a linha existente como duplicada
                        for row_cells in ws.iter_rows(min_row=4):
                            if tuple(cell.value for cell in row_cells) == tuple(novo_item):
                                for cell in row_cells:
                                    cell.fill = amarelo_duplicado
                                break
                        break

                # Determina a próxima linha disponível
                proxima_linha = ws.max_row + 1

                # Insere o novo item
                for col_num, value in enumerate(novo_item, 1):
                    cell = ws.cell(row=proxima_linha, column=col_num, value=value)
                    cell.font = fonte_normal
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

                    # Aplica formatação específica por coluna
                    if col_num == 1:  # N. Nota - Inteiro
                        cell.number_format = formato_inteiro
                        if isinstance(value, str) and value.isdigit():
                            cell.value = int(value)
                    elif col_num == 2:  # Data - Formato data
                        cell.number_format = formato_data
                    elif col_num == 3:  # Linha da PO - Inteiro
                        cell.number_format = formato_inteiro
                        if isinstance(value, str) and value.isdigit():
                            cell.value = int(value)
                    elif col_num == 4:  # Descrição - Texto
                        cell.number_format = formato_texto
                    elif col_num == 5:  # Valor - Moeda com 2 decimais
                        cell.number_format = formato_moeda
                        if isinstance(value, str):
                            # Remove possíveis pontos de milhar e converte vírgula para ponto
                            value = value.replace('.', '').replace(',', '.')
                            try:
                                cell.value = float(value)
                            except ValueError:
                                pass

                    # Formatação de cor de fundo
                    if duplicata:
                        cell.fill = vermelho_novo
                    else:
                        if proxima_linha % 2 == 0:  # Linhas pares - branco
                            pass  # Sem preenchimento
                        else:  # Linhas ímpares - azul claro
                            cell.fill = azul_linha

                # Ajusta altura da linha para 40 pixels
                ws.row_dimensions[proxima_linha].height = 40

            # Aplicar bordas às células já existentes (linhas 1-3)
            for row in ws.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    cell.border = thin_border

        # Reordenar as planilhas por número de PO
        wb._sheets.sort(key=lambda ws: int(ws.title.split('_')[1]) if ws.title.startswith('PO_') else 0)

        # Salvar o arquivo
        wb.save(caminho_completo)
        return True

    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return False